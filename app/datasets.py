from __future__ import annotations

import math
import os
import shutil
import tempfile
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Callable, Mapping

from openpyxl import load_workbook

from .config import PROJECT_ROOT
from .data import FEATURE_NAMES


DATASET_DIR = PROJECT_ROOT / "datasets_ready"


def next_quarter(period: str) -> str:
    if len(period) != 6 or period[4] != "Q" or not period[:4].isdigit() or period[5] not in "1234":
        raise ValueError(f"Некорректный квартал: {period}")
    year, quarter = int(period[:4]), int(period[5])
    return f"{year + 1}Q1" if quarter == 4 else f"{year}Q{quarter + 1}"


class DatasetStore:
    """Безопасное чтение и атомарное изменение XLSX из datasets_ready."""

    def __init__(self, directory: Path | str = DATASET_DIR):
        self.directory = Path(directory).resolve()

    def _resolve(self, name: str) -> Path:
        if Path(name).name != name or not name.lower().endswith(".xlsx"):
            raise ValueError("Можно выбирать только XLSX-файлы из datasets_ready")
        path = (self.directory / name).resolve()
        if path.parent != self.directory or not path.is_file():
            raise FileNotFoundError(f"Датасет не найден: {name}")
        return path

    @staticmethod
    def _sheet(path: Path):
        workbook = load_workbook(path, read_only=True, data_only=True)
        if "Лист1" not in workbook.sheetnames:
            workbook.close()
            raise ValueError("В XLSX отсутствует лист «Лист1»")
        return workbook, workbook["Лист1"]

    def names(self) -> list[str]:
        return sorted(path.name for path in self.directory.glob("*.xlsx") if path.is_file() and not path.name.startswith("."))

    def path(self, name: str) -> Path:
        return self._resolve(name)

    def summary(self, name: str) -> dict[str, object]:
        path = self._resolve(name)
        workbook, sheet = self._sheet(path)
        try:
            rows = max(sheet.max_row - 2, 0)
            first_period = str(sheet.cell(3, 1).value) if rows else None
            last_period = str(sheet.cell(sheet.max_row, 1).value) if rows else None
            return {
                "name": name,
                "rows": rows,
                "first_period": first_period,
                "last_period": last_period,
                "size_bytes": path.stat().st_size,
            }
        finally:
            workbook.close()

    def catalog(self, active_name: str) -> dict[str, object]:
        return {
            "directory": "datasets_ready",
            "active": active_name,
            "datasets": [self.summary(name) | {"active": name == active_name} for name in self.names()],
        }

    def read(self, name: str) -> dict[str, object]:
        path = self._resolve(name)
        workbook, sheet = self._sheet(path)
        try:
            groups: list[str] = []
            current_group = ""
            for column in range(6, 37):
                value = sheet.cell(1, column).value
                if value not in (None, ""):
                    current_group = str(value).strip()
                groups.append(current_group)
            columns = [
                {
                    "id": feature,
                    "group": group,
                    "label": str(sheet.cell(2, column).value).strip(),
                }
                for feature, group, column in zip(FEATURE_NAMES, groups, range(6, 37), strict=True)
            ]
            rows = []
            for row in range(3, sheet.max_row + 1):
                period = str(sheet.cell(row, 1).value)
                rows.append(
                    {
                        "period": period,
                        "values": {
                            feature: float(sheet.cell(row, column).value)
                            for feature, column in zip(FEATURE_NAMES, range(6, 37), strict=True)
                        },
                    }
                )
            return {
                **self.summary(name),
                "columns": columns,
                "rows_data": rows,
                "next_period": next_quarter(rows[-1]["period"]),
            }
        finally:
            workbook.close()

    @staticmethod
    def _validated_values(values: Mapping[str, float], *, require_all: bool) -> dict[str, float]:
        unknown = set(values) - set(FEATURE_NAMES)
        if unknown:
            raise ValueError(f"Неизвестные показатели: {', '.join(sorted(unknown))}")
        if require_all and set(values) != set(FEATURE_NAMES):
            missing = set(FEATURE_NAMES) - set(values)
            raise ValueError(f"Нужно заполнить все 31 показателя; отсутствуют: {', '.join(sorted(missing))}")
        normalized = {name: float(value) for name, value in values.items()}
        if any(not math.isfinite(value) for value in normalized.values()):
            raise ValueError("Значения должны быть конечными числами")
        if any(value < 0 for value in normalized.values()):
            raise ValueError("Для показателей датасета отрицательные значения не допускаются")
        return normalized

    def _atomic_edit(
        self,
        name: str,
        edit: Callable[[object], str],
        validator: Callable[[Path], object],
    ) -> str:
        target = self._resolve(name)
        descriptor, temp_name = tempfile.mkstemp(prefix=f".{target.stem}-", suffix=".xlsx", dir=self.directory)
        os.close(descriptor)
        Path(temp_name).unlink(missing_ok=True)
        try:
            shutil.copy2(target, temp_name)
            workbook = load_workbook(temp_name)
            try:
                result = edit(workbook)
                workbook.save(temp_name)
            finally:
                workbook.close()
            validator(Path(temp_name))
            Path(temp_name).replace(target)
            return result
        finally:
            Path(temp_name).unlink(missing_ok=True)

    def update_row(
        self,
        name: str,
        period: str,
        values: Mapping[str, float],
        validator: Callable[[Path], object],
    ) -> str:
        normalized = self._validated_values(values, require_all=True)

        def edit(workbook) -> str:
            sheet = workbook["Лист1"]
            row_number = next(
                (row for row in range(3, sheet.max_row + 1) if str(sheet.cell(row, 1).value) == period),
                None,
            )
            if row_number is None:
                raise ValueError(f"Период {period} не найден")
            for feature, column in zip(FEATURE_NAMES, range(6, 37), strict=True):
                sheet.cell(row_number, column, normalized[feature])
            return period

        return self._atomic_edit(name, edit, validator)

    def append_row(
        self,
        name: str,
        values: Mapping[str, float],
        validator: Callable[[Path], object],
    ) -> str:
        normalized = self._validated_values(values, require_all=True)

        def edit(workbook) -> str:
            sheet = workbook["Лист1"]
            previous_row = sheet.max_row
            row_number = previous_row + 1
            period = next_quarter(str(sheet.cell(previous_row, 1).value))
            year, quarter = int(period[:4]), int(period[5])
            for column in range(1, sheet.max_column + 1):
                source, target = sheet.cell(previous_row, column), sheet.cell(row_number, column)
                if source.has_style:
                    target._style = copy(source._style)
                target.number_format = source.number_format
                target.font = copy(source.font)
                target.fill = copy(source.fill)
                target.border = copy(source.border)
                target.alignment = copy(source.alignment)
                target.protection = copy(source.protection)
            sheet.row_dimensions[row_number].height = sheet.row_dimensions[previous_row].height
            metadata = (period, datetime(year, 1 + (quarter - 1) * 3, 1), year, quarter, row_number - 2)
            for column, value in enumerate(metadata, start=1):
                sheet.cell(row_number, column, value)
            for feature, column in zip(FEATURE_NAMES, range(6, 37), strict=True):
                sheet.cell(row_number, column, normalized[feature])
            sheet.cell(row_number, 37, 0.0)
            return period

        return self._atomic_edit(name, edit, validator)
